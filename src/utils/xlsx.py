import datetime
import re

import pandas as pd  # type: ignore
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

_RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


def _sanitize_for_excel(value: str) -> str:
    """Strip ANSI escape codes and other control characters illegal in Excel cells."""
    # Remove ANSI escape sequences
    value = re.sub(r"\x1b\[[0-9;]*[a-zA-Z]", "", value)
    # Remove remaining illegal control characters (openpyxl disallows 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F)
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    return value


# --------------------------------- append_to_failed_downloads ---------------------------------
def append_to_failed_downloads(file, url, reason):
    try:
        failed_df = pd.read_excel(file, sheet_name="failedDownloads")
    except (FileNotFoundError, ValueError):
        failed_df = pd.DataFrame(columns=["Date", "URL", "Reason"])

    new_row = pd.DataFrame([{
        "Date": datetime.date.today().strftime("%m/%d/%Y"),
        "URL": url,
        "Reason": _sanitize_for_excel(str(reason)),
    }])

    updated_df = pd.concat([failed_df, new_row], ignore_index=True)

    with pd.ExcelWriter(
        file, mode="a", if_sheet_exists="replace", engine="openpyxl"
    ) as writer:
        updated_df.to_excel(writer, sheet_name="failedDownloads", index=False)


# --------------------------------- deduplicate_failed_downloads ---------------------------------
def deduplicate_failed_downloads(file):
    try:
        failed_df = pd.read_excel(file, sheet_name="failedDownloads")
    except (FileNotFoundError, ValueError):
        return

    deduped_df = failed_df.drop_duplicates(subset=["URL"], keep="last")

    if len(deduped_df) == len(failed_df):
        return

    with pd.ExcelWriter(
        file, mode="a", if_sheet_exists="replace", engine="openpyxl"
    ) as writer:
        deduped_df.to_excel(writer, sheet_name="failedDownloads", index=False)


# --------------------------------- mark_url_red ---------------------------------
def mark_url_red(file, url, sheet_name):
    """Color the row matching `url` red in the given sheet."""
    wb = load_workbook(file)
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # Find the URL column index from the header row
    header = next(ws.iter_rows(min_row=1, max_row=1))
    url_col = next(
        (cell.column for cell in header if str(cell.value or "").strip().lower() == "url"),
        None,
    )
    if url_col is None:
        return

    for row in ws.iter_rows(min_row=2):
        if row[url_col - 1].value == url:
            for cell in row:
                cell.fill = _RED_FILL
            break

    wb.save(file)


# --------------------------------- is_already_downloaded ---------------------------------
def is_already_downloaded(file, title, uploader):
    try:
        past_df = pd.read_excel(file, sheet_name="pastDownloads")
    except (FileNotFoundError, ValueError):
        return False

    if "Title" not in past_df.columns or "Uploader" not in past_df.columns:
        return False

    match = past_df[
        (past_df["Title"].str.strip().str.lower() == title.strip().lower()) &
        (past_df["Uploader"].str.strip().str.lower() == uploader.strip().lower())
    ]
    return not match.empty


# --------------------------------- append_to_past_downloads ---------------------------------
def append_to_past_downloads(file, url, title, uploader):
    # Try to load the existing 'pastDownloads' sheet
    try:
        past_df = pd.read_excel(file, sheet_name="pastDownloads")
    except (FileNotFoundError, ValueError):
        # If the file or sheet doesn't exist, start a new DataFrame
        past_df = pd.DataFrame(columns=["Date Downloaded", "URL", "Title", "Uploader"])

    # Create the new row
    new_row = pd.DataFrame([{
        "Date Downloaded": datetime.date.today().strftime("%m/%d/%Y"),
        "URL": url,
        "Title": title,
        "Uploader": uploader,
    }])

    # Append and write back
    updated_df = pd.concat([past_df, new_row], ignore_index=True)

    with pd.ExcelWriter(
        file, mode="a", if_sheet_exists="replace", engine="openpyxl"
    ) as writer:
        updated_df.to_excel(writer, sheet_name="pastDownloads", index=False)
