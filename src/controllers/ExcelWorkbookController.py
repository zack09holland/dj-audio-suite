import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelWorkbookController:
    def __init__(self, file_path):
        """
        Initialize the ExcelHandler object.

        Parameters:
        file_path (str): The path to the Excel file.
        """
        self.file_path = file_path
        self.workbook = None

    def initialize_workbook(self):
        """
        Load the workbook from the file path. If the file does not exist, create a new workbook.
        """
        try:
            self.workbook = load_workbook(self.file_path)

            print(f"Workbook '{self.file_path}' initialized.")
        except FileNotFoundError:
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)
            print(f"File not found. Created a new workbook.")

    def get_workbook(self):
        """
        Reads the workbook from the file path and creates a pandas dataframe of it
        """
        try:
            # self.workbook = load_workbook(self.file_path)
            xls = pd.ExcelFile(self.file_path)
            print(f"Workbook dataframe acquired.")
            return xls
        except FileNotFoundError:
            print(f"File not found")

    def save_workbook(self):
        """
        Save the workbook to the specified file path.
        """
        if not self.workbook:
            raise Exception(
                "No workbook loaded or created. Call load_workbook() first."
            )
        self.workbook.save(self.file_path)
        print(f"Workbook saved as '{self.file_path}'.")

    def reset_sheet(self, sheet_name):
        """
        Resets a sheet by removing it and creating a new one with the same name.

        Parameters:
        sheet_name (str): Name of the sheet to reset.
        """
        if not self.workbook:
            print("No workbook loaded. Call 'load_workbook' first.")
            return

        if sheet_name in self.workbook.sheetnames:
            # Remove the existing sheet
            del self.workbook[sheet_name]
            print(f"Sheet '{sheet_name}' has been reset.")

        # Create a new, empty sheet with the same name
        self.workbook.create_sheet(sheet_name)
        print(f"New empty sheet '{sheet_name}' created.")

    def add_df_to_workbook(self, df, sheet_name):
        """
        Adds a DataFrame to a workbook sheet. Creates the sheet if it doesn't exist
        and ensures the headers match the expected format.

        Parameters:
        df (pd.DataFrame): DataFrame containing the data to add.
        sheet_name (str): Name of the sheet to add the data to.
        """
        if not self.workbook:
            print("No workbook loaded. Call 'load_workbook' first.")
            return

        print(f"Adding data to sheet: {sheet_name}")
        try:
            if df.empty:
                print(f"DataFrame is empty. No data added to sheet '{sheet_name}'.")
                return

            # Check if the sheet exists; if not, create it
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                print(f"Sheet '{sheet_name}' exists. Appending data.")
            else:
                sheet = self.workbook.create_sheet(sheet_name)
                print(f"Created new sheet: {sheet_name}.")

            # Append DataFrame rows to the sheet
            rows_before = sheet.max_row
            for row in dataframe_to_rows(df, index=False, header=True):
                sheet.append(row)
            rows_added = sheet.max_row - rows_before
            print(f"Added {rows_added} rows to sheet '{sheet_name}'.")

        except Exception as e:
            print(f"Error adding data to sheet '{sheet_name}': {e}")

    def list_sheets(self):
        """
        List all sheet names in the workbook.

        Returns:
        list: A list of sheet names.
        """
        if not self.workbook:
            raise Exception(
                "No workbook loaded or created. Call load_workbook() first."
            )
        return self.workbook.sheetnames

    def highlight_rows(self, row):
        # If update type is a withdrawal, color it a light red
        if row["Update Type"] == "W":
            return ["background-color: #FFC7CE"] * len(row)
        if row["Update Type"] == "N":
            return ["background-color: #C6EFCE"] * len(row)
        else:
            return [""] * len(row)

    def highlight_row(row):
        print(row)
        if row.Name != "Name":
            print("row.Name")
            if row.Count < 40:
                return pd.Series("background-color: red", row.index)
            else:
                return pd.Series("background-color: green", row.index)
        else:
            return pd.Series("", row.index)


if __name__ == "__main__":
    # An example DataFrame
    data = {
        "Name": ["Alice", "Bob", "Charlie"],
        "Age": [25, 30, 35],
        "Update Type": ["N", "W", "R"],
        "City": ["New York", "San Francisco", "Los Angeles"],
    }
    df = pd.DataFrame(data)

    # Initialize ExcelHandler with the file path
    excel_handler = ExcelWorkbookController("example.xlsx")

    # Load or create the workbook
    excel_handler.initialize_workbook()
    df.style.apply(excel_handler.highlight_row, axis=1)

    # Add the DataFrame to a new sheet
    excel_handler.add_df_to_workbook(df, sheet_name="Sheet1")

    # Save the workbook
    excel_handler.save_workbook()

    # List all sheets
    print("Available sheets:", excel_handler.list_sheets())